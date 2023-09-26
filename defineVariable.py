# Reading Excel
EXCEL_FILE = 'formb4b9.xlsx'
from openpyxl import load_workbook
book = load_workbook(EXCEL_FILE) 
sheet = book.active

# Link 
test = 'https://google.com'
formb4b9='https://docs.google.com/forms/d/e/1FAIpQLSdhmCXo_17SP0oDj82XmEvp8RFJyofVL8J5Iqss73EZojTdXQ/viewform'

# Constant - Teacher Detail
TEACHER_NAME = sheet['G2'].value
TEACHER_CODE = sheet['H2'].value
# Day Report 
DAY_REPORT = sheet['K2'].value

# Report code 
LESSON = sheet['I2'].value
def getReportCode(lesson):
    match lesson:
        case 4:
            return 3
        case 9:
            return 4
        case 12:
            return 5
        case 'Upsale':
            return 6
        case 'upsale':
            return 6
        case default:
            return 3
REPORT_CODE =  getReportCode(LESSON)

# Text Field 
XPATH_TeacherName = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[1]/div/div/div[2]/div/div[1]/div/div[1]/input"
XPATH_TeacherCode = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[2]/div/div/div[2]/div/div[1]/div/div[1]/input"
XPATH_DROPDOWN_CENTER_CODE = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[3]/div/div/div[2]/div/div[1]/div[1]/div[1]"
XPATH_DROPDOWN_REPORT_CODE = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[7]/div/div/div[2]/div/div[1]/div[1]/div[1]"
XPATH_DATE_REPORT = "/html/body/div/div[2]/form/div[2]/div/div[2]/div[6]/div/div/div[2]/div/div/div[2]/div[1]/div/div[1]/input"


# Next Button 
XPATH_NEXTBUTTON = '/html/body/div/div[2]/form/div[2]/div/div[3]/div[1]/div[1]/div/span'

# Submit Button
XPATH_SUBMITBUTTON = ''

# Append data to raw_data
raw_data = []
for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
    raw_data.append(row)